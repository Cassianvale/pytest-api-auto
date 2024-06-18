#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
import shutil
import ast
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from common.setting import ensure_path_sep
from utils.read_files_tools.get_all_files_path import get_all_files
from utils.notify.wechat_send import WeChatSend
from utils.other_tools.allure_data.allure_report_data import AllureFileClean


class ErrorTestCase:
    """ 收集错误的用例 """

    def __init__(self):
        self.test_case_path = ensure_path_sep("\\report\\html\\data\\test-cases\\")

    def get_error_case_data(self):
        """
        收集所有失败用例的数据
        @return:
        """
        path = get_all_files(self.test_case_path)
        files = []
        for i in path:
            with open(i, 'r', encoding='utf-8') as file:
                date = json.load(file)
                # 收集执行失败的用例数据
                if date['status'] == 'failed' or date['status'] == 'broken':
                    files.append(date)
        return files

    @classmethod
    def get_case_name(cls, test_case):
        """
        收集测试用例名称
        @return:
        """
        name = test_case['name'].split('[')
        case_name = name[1][:-1]
        return case_name

    @classmethod
    def get_parameters(cls, test_case):
        """
        获取allure报告中的 parameters 参数内容, 请求前的数据
        用于兼容用例执行异常，未发送请求导致的情况
        @return:
        """
        parameters = test_case.get('parameters', [{}])[0].get('value', '{}')
        return ast.literal_eval(parameters)

    @classmethod
    def get_test_stage(cls, test_case):
        """
        获取allure报告中请求后的数据
        @return:
        """
        return test_case.get('testStage', {}).get('steps', [])

    def get_case_url(self, test_case):
        """
        获取测试用例的 url
        @param test_case:
        @return:
        """
        test_stage = test_case.get('testStage', {})
        # 判断用例步骤中的数据是否异常
        if test_stage.get('status') == 'broken':
            # 如果异常状态下，则获取请求前的数据
            _url = self.get_parameters(test_case).get('url', 'URL not found')
        else:
            # 否则拿请求步骤的数据，因为如果涉及到依赖，会获取多组，因此我们只取最后一组数据内容
            steps = test_stage.get('steps', [])
            _url = steps[-7]['name'][7:] if len(steps) > 6 else 'URL not found'
        return _url

    def get_method(self, test_case):
        """
        获取用例中的请求方式
        @param test_case:
        @return:
        """
        test_stage = test_case.get('testStage', {})
        if test_stage.get('status') == 'broken':
            _method = self.get_parameters(test_case).get('method', 'Method not found')
        else:
            steps = test_stage.get('steps', [])
            _method = steps[-6]['name'][6:] if len(steps) > 5 else 'Method not found'
        return _method

    def get_headers(self, test_case):
        """
        获取用例中的请求头
        @return:
        """
        test_stage = test_case.get('testStage', {})
        if test_stage.get('status') == 'broken':
            _headers = self.get_parameters(test_case).get('headers', 'Headers not found')
        else:
            # 如果用例请求成功，则从allure附件中获取请求头部信息
            steps = test_stage.get('steps', [])
            if len(steps) > 4 and 'attachments' in steps[-5]:
                _headers_attachment = steps[-5]['attachments'][0]['source']
                path = ensure_path_sep("\\report\\html\\data\\attachments\\" + _headers_attachment)
                try:
                    with open(path, 'r', encoding='utf-8') as file:
                        _headers = json.load(file)
                except FileNotFoundError:
                    _headers = 'Headers file not found'
            else:
                _headers = 'Headers not found'
        return _headers

    def get_case_data(self, test_case):
        """
        获取用例内容
        @return:
        """
        test_stage = test_case.get('testStage', {})
        if test_stage.get('status') == 'broken':
            _case_data = self.get_parameters(test_case).get('data', 'Data not found')
        else:
            steps = test_stage.get('steps', [])
            if len(steps) > 3 and 'attachments' in steps[-4]:
                _case_data_attachments = steps[-4]['attachments'][0]['source']
                path = ensure_path_sep("\\report\\html\\data\\attachments\\" + _case_data_attachments)
                try:
                    with open(path, 'r', encoding='utf-8') as file:
                        _case_data = json.load(file)
                except FileNotFoundError:
                    _case_data = 'Case data file not found'
            else:
                _case_data = 'Data not found'
        return _case_data

    def get_dependence_case(self, test_case):
        """
        获取依赖用例
        @param test_case:
        @return:
        """
        return self.get_parameters(test_case).get('dependence_case_data', 'Dependence case not found')

    def get_sql(self, test_case):
        """
        获取 sql 数据
        @param test_case:
        @return:
        """
        return self.get_parameters(test_case).get('sql', 'SQL not found')

    def get_assert(self, test_case):
        """
        获取断言数据
        @param test_case:
        @return:
        """
        return self.get_parameters(test_case).get('assert_data', 'Assert data not found')

    def get_request_type(self, test_case):
        """
        获取用例的请求类型
        @param test_case:
        @return:
        """
        return self.get_parameters(test_case).get('requestType', 'Request type not found')

    @classmethod
    def get_response(cls, test_case):
        """
        获取响应内容的数据
        @param test_case:
        @return:
        """
        test_stage = test_case.get('testStage', {})
        if test_stage.get('status') == 'broken':
            return test_stage.get('statusMessage', 'Response not found')
        else:
            steps = test_stage.get('steps', [])
            if not steps:
                return 'Response not found'
            try:
                res_data_attachments = steps[-1].get('attachments', [])[0]['source']
                path = ensure_path_sep("\\report\\html\\data\\attachments\\" + res_data_attachments)
                with open(path, 'r', encoding='utf-8') as file:
                    return json.load(file)
            except FileNotFoundError:
                return 'Response file not found'
            except (KeyError, IndexError):
                return 'Response not found'

    @classmethod
    def get_case_time(cls, test_case):
        """
        获取用例运行时长
        @param test_case:
        @return:
        """
        return str(test_case.get('time', {}).get('duration', '0')) + "ms"

    @classmethod
    def get_uid(cls, test_case):
        """
        获取 allure 报告中的 uid
        @param test_case:
        @return:
        """
        return test_case.get('uid', 'UID not found')


class ErrorCaseExcel:
    """ 收集运行失败的用例，整理成excel报告 """

    def __init__(self):
        self.case_data = ErrorTestCase()
        self._file_path = ensure_path_sep("\\Files\\test_data\\自动化异常测试用例.xlsx")
        _excel_template = ensure_path_sep("\\utils\\other_tools\\allure_data\\自动化异常测试用例.xlsx")
        shutil.copyfile(src=_excel_template, dst=self._file_path)

    def write_case(self):
        _data = self.case_data.get_error_case_data()
        if len(_data) > 0:
            data = []
            for case in _data:
                row = {
                    'UID': self.case_data.get_uid(case),
                    '用例名称': self.case_data.get_case_name(case),
                    'URL': self.case_data.get_case_url(case),
                    '请求方式': self.case_data.get_method(case),
                    '请求类型': self.case_data.get_request_type(case),
                    '请求头': self.case_data.get_headers(case),
                    '请求内容': self.case_data.get_case_data(case),
                    '依赖用例': self.case_data.get_dependence_case(case),
                    '断言内容': self.case_data.get_assert(case),
                    'SQL': self.case_data.get_sql(case),
                    '运行时长': self.case_data.get_case_time(case),
                    '响应内容': self.case_data.get_response(case)
                }
                data.append(row)

            df = pd.DataFrame(data)
            df.to_excel(self._file_path, sheet_name='异常用例', index=False)

            # 使用 openpyxl 加载生成的 Excel 文件，并设置样式
            self.format_excel()

            WeChatSend(AllureFileClean().get_case_count()).send_file_msg(self._file_path)

    def format_excel(self):
        workbook = load_workbook(self._file_path)
        sheet = workbook['异常用例']

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="548235")
        header_alignment = Alignment(vertical="center", horizontal="center")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 表头高度
        sheet.row_dimensions[1].height = 20

        # 列宽度
        column_widths = {
            'A': 15,  # UID
            'B': 20,  # 用例名称
            'C': 25,  # URL
            'D': 10,  # 请求方式
            'E': 10,  # 请求类型
            'F': 30,  # 请求头
            'G': 30,  # 请求内容
            'H': 30,  # 依赖用例
            'I': 30,  # 断言内容
            'J': 30,  # SQL
            'K': 10,  # 运行时长
            'L': 30  # 响应内容
        }

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # 设置自动换行、行高以及顶端对齐和左对齐
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                # 设置行高
                sheet.row_dimensions[cell.row].height = 50

        workbook.save(self._file_path)
        workbook.close()


if __name__ == "__main__":
    excel = ErrorCaseExcel()
    excel.write_case()
